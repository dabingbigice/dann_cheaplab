import cv2
import numpy as np
import os
import pandas as pd
from datetime import datetime
import math
import matplotlib.pyplot as plt
from openpyxl.styles import PatternFill


def ellipse_perimeter(a, b):
    return math.pi * (3 * (a + b) - math.sqrt((3 * a + b) * (a + 3 * b)))


def batch_folder_predict(input_folder, output_dir="batch_result_visual",
                         extensions=('jpg', 'png', 'jpeg', 'bmp', 'tiff'),
                         excel_output="板栗特征统计.xlsx", px_area=1.0, px_long=1.0, h=1.0):
    """
    批量处理文件夹中的所有图片并输出特征到Excel
    :param input_folder: 输入文件夹路径
    :param output_dir: 输出结果文件夹路径
    :param extensions: 支持的图片格式扩展名
    :param excel_output: Excel输出文件名
    :param px_area: 像素面积标定系数
    :param px_long: 像素长度标定系数
    :param h: 厚度参数
    :return: 所有图片的处理结果列表
    """
    os.makedirs(output_dir, exist_ok=True)

    # 获取所有图片文件
    image_files = []
    for filename in os.listdir(input_folder):
        if filename.lower().endswith(extensions):
            image_files.append(filename)

    if not image_files:
        print(f"在文件夹 {input_folder} 中未找到支持的图片文件")
        return []

    print(f"找到 {len(image_files)} 个图片文件，开始批量处理...")

    all_results = []
    features_list = []  # 存储所有特征数据的列表

    # 遍历处理每个图片文件
    for i, filename in enumerate(image_files):
        print(f"\n正在处理第 {i + 1}/{len(image_files)} 个文件: {filename}")
        image_path = os.path.join(input_folder, filename)

        result, a, b, perimeter, error, ellipse_params = single_file_predict(
            image_path, output_dir, px_area, px_long, h)

        # 记录文件名到结果中
        result['filename'] = filename
        all_results.append({
            'result': result,
            'a': a,
            'b': b,
            'perimeter': perimeter,
            'error': error,
            'ellipse_params': ellipse_params
        })

        # 如果处理成功，提取特征并添加到列表中
        if result['status'] == 'success' and a is not None and b is not None:
            features = extract_walnut_features(a, b, result['actual_area'], px_area, px_long, h, filename)
            features_list.append(features)
            print(f"  ✓ 成功处理: 误差 {error:.2f}%")
        else:
            print(f"  ✗ 处理失败: {result['message']}")

    # 如果有成功处理的结果，生成Excel文件
    if features_list:
        export_features_to_excel(features_list, excel_output)
        print(f"\n✓ 特征数据已导出到: {excel_output}")

    # 生成批量处理总结报告
    generate_batch_report(all_results, output_dir)

    return all_results, features_list


def extract_walnut_features(a, b, actual_area, px_area, px_long, h, filename):
    """
    提取板栗仁的各种特征参数
    """
    # 板栗仁标定面积
    hutao_area = px_area * actual_area

    # 拟合椭圆长半轴标定长
    hutao_a = a * px_long

    # 拟合椭圆短半轴标定长
    hutao_b = b * px_long

    # ab算术均值
    arithmetic_a_b_avg = (hutao_a + hutao_b) / 2

    # ab几何均值
    geometry_a_b_avg = (hutao_a * hutao_b) ** (1 / 2)

    # 长轴全长、宽轴全长、厚度
    L = hutao_a * 2
    W = hutao_b * 2


    # 表面积和体积计算
    vops = ((4 / 6) * np.pi) * L * W * W
    voos = ((4 / 6) * np.pi) * L * L * W

    # 椭圆周长（标定后）
    hutao_perimeter = ellipse_perimeter(hutao_a, hutao_b)

    # 将所有特征打包成字典[6,7](@ref)
    features = {
        '文件名': filename,
        '像素面积': actual_area,
        '标定面积': round(hutao_area, 4),
        '长半轴(a)': round(hutao_a, 4),
        '短半轴(b)': round(hutao_b, 4),
        '长轴全长(L)': round(L, 4),
        '宽轴全长(W)': round(W, 4),
        '椭圆周长': round(hutao_perimeter, 4),
        'ab算术均值': round(arithmetic_a_b_avg, 4),
        'ab几何均值': round(geometry_a_b_avg, 4),
        '体积(vops)': round(vops, 4),
        '体积(voos)': round(voos, 4),
        '像素标定系数(面积)': px_area,
        '像素标定系数(长度)': px_long
    }

    return features


def export_features_to_excel(features_list, excel_output):
    """
    将特征数据导出到Excel文件[6,7](@ref)
    """
    # 创建DataFrame
    df = pd.DataFrame(features_list)

    # 设置文件名作为索引
    df.set_index('文件名', inplace=True)

    # 导出到Excel[7,8](@ref)
    try:
        df.to_excel(excel_output, sheet_name='板栗特征统计',
                    engine='openpyxl', index=True)

        # 添加格式美化（可选）
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter

            # 加载工作簿进行格式设置
            workbook = load_workbook(excel_output)
            worksheet = workbook.active

            # 设置标题行格式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            for col in range(1, len(df.columns) + 2):  # +2 因为包含索引列
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            # 自动调整列宽
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

            workbook.save(excel_output)
        except ImportError:
            print("注意: openpyxl样式设置不可用，导出基本Excel文件")

    except Exception as e:
        print(f"导出Excel文件时出错: {e}")
        # 备用方案：导出为CSV
        csv_output = excel_output.replace('.xlsx', '.csv')
        df.to_csv(csv_output, index=True)
        print(f"特征数据已导出为CSV: {csv_output}")


def single_file_predict(image_path, output_dir="result_visual", px_area=1.0, px_long=1.0, h=1.0):
    os.makedirs(output_dir, exist_ok=True)

    result = {
        'status': 'success',
        'message': '',
        'actual_area': 0,
        'predicted_area': 0,
        'error_percent': 0,
        'axes': (0, 0),
        'vis_path': '',
        'ellipse_params': {}
    }

    try:
        mask = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
        if mask is None:
            raise ValueError("无法读取图像文件")

        kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (7, 7))
        closed = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel, iterations=2)

        actual_area = np.count_nonzero(closed)
        result['actual_area'] = actual_area
        if actual_area < 100:
            raise ValueError(f"非0像素过少（仅{actual_area}个），可能是分割失败或噪声")

        contours, _ = cv2.findContours(closed, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        if not contours:
            raise ValueError("未检测到有效轮廓")

        largest_contour = max(contours, key=cv2.contourArea)
        if not is_valid_contour(largest_contour):
            raise ValueError("无效轮廓")

        ellipse = cv2.fitEllipseDirect(largest_contour)
        (center, (width, height), angle) = ellipse

        ellipse_adj = (center, (width, height), angle)
        b = max(width, height) / 2.0
        a = min(width, height) / 2.0
        ellipse_area = np.pi * a * b

        major_axis_len = round(a * 2, 1)
        minor_axis_len = round(b * 2, 1)
        result['predicted_area'] = ellipse_area
        result['error_percent'] = abs((ellipse_area - actual_area) / actual_area) * 100
        result['axes'] = (major_axis_len, minor_axis_len)
        result['ellipse_params'] = {
            "center": (round(center[0], 1), round(center[1], 1)),
            "major_axis": major_axis_len,
            "minor_axis": minor_axis_len,
            "angle": round(angle, 1)
        }

        print(f"处理 {os.path.basename(image_path)}:")
        print(f"  实际面积（非0像素数）: {actual_area} px²")
        print(f"  椭圆拟合面积: {ellipse_area:.1f} px²")
        print(f"  面积误差: {result['error_percent']:.2f}%")
        print(
            f"  椭圆参数: 长轴a={major_axis_len}px, 短轴b={minor_axis_len}px, 中心{result['ellipse_params']['center']}, 角度{result['ellipse_params']['angle']}°")

        base_name = os.path.splitext(os.path.basename(image_path))[0]
        vis_path = os.path.join(output_dir, f"result_{base_name}.png")
        visualize_result(mask, closed, ellipse, ellipse_adj, actual_area, vis_path, a, b, angle, major_axis_len,
                         minor_axis_len, result['error_percent'], ellipse_area)
        result['vis_path'] = vis_path

        return result, a, b, ellipse_perimeter(a, b), result['error_percent'], result['ellipse_params']

    except Exception as e:
        result['status'] = 'error'
        result['message'] = str(e)
        print(f"处理失败: {e}")
        return result, None, None, None, None, {}


def generate_batch_report(all_results, output_dir):
    """生成批量处理总结报告"""
    report_path = os.path.join(output_dir, "batch_processing_report.txt")

    successful_results = [r for r in all_results if r['result']['status'] == 'success']
    failed_results = [r for r in all_results if r['result']['status'] == 'error']

    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("=" * 60 + "\n")
        f.write("批量椭圆拟合处理报告\n")
        f.write("=" * 60 + "\n")
        f.write(f"处理时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"总文件数: {len(all_results)}\n")
        f.write(f"成功处理: {len(successful_results)}\n")
        f.write(f"处理失败: {len(failed_results)}\n")
        f.write("=" * 60 + "\n\n")

        if successful_results:
            f.write("成功处理文件详情:\n")
            f.write("-" * 40 + "\n")

            # 计算统计信息
            errors = [r['error'] for r in successful_results]
            areas = [r['result']['actual_area'] for r in successful_results]

            for i, result_data in enumerate(successful_results):
                r = result_data['result']
                f.write(f"{i + 1}. {r['filename']}:\n")
                f.write(f"   实际面积: {r['actual_area']} px²\n")
                f.write(f"   预测面积: {r['predicted_area']:.1f} px²\n")
                f.write(f"   误差: {r['error_percent']:.2f}%\n")
                f.write(f"   椭圆参数: 长轴{r['axes'][0]}px, 短轴{r['axes'][1]}px\n")
                f.write(f"   可视化文件: {os.path.basename(r['vis_path'])}\n\n")

            f.write("统计信息:\n")
            f.write(f"平均误差: {np.mean(errors):.2f}%\n")
            f.write(f"最大误差: {max(errors):.2f}%\n")
            f.write(f"最小误差: {min(errors):.2f}%\n")
            f.write(f"平均面积: {np.mean(areas):.0f} px²\n")

        if failed_results:
            f.write("\n处理失败文件详情:\n")
            f.write("-" * 40 + "\n")
            for i, result_data in enumerate(failed_results):
                r = result_data['result']
                f.write(f"{i + 1}. {r['filename']}: {r['message']}\n")

    print(f"\n批量处理完成！详细报告已保存至: {report_path}")

    # 在控制台显示简要总结
    print("\n" + "=" * 50)
    print("批量处理总结:")
    print(f"总文件数: {len(all_results)}")
    print(f"成功处理: {len(successful_results)}")
    print(f"处理失败: {len(failed_results)}")

    if successful_results:
        errors = [r['error'] for r in successful_results]
        print(f"平均误差: {np.mean(errors):.2f}%")
        print(f"误差范围: {min(errors):.2f}% - {max(errors):.2f}%")


def is_center_region(x: float, y: float) -> bool:
    width, height = 512, 468
    center_size = int(width * math.sqrt(0.6))
    margin = (width - center_size) // 2
    x_min, x_max = margin, width - margin - 1
    y_min, y_max = margin, height - margin - 1
    return (x_min <= x <= x_max) and (y_min <= y <= y_max)


def is_valid_contour(cnt):
    area = cv2.contourArea(cnt)
    if area < 100 or len(cnt) < 15:
        return False
    hull = cv2.convexHull(cnt)
    hull_area = cv2.contourArea(hull)
    if hull_area == 0:
        return False
    return (area / hull_area) > 0.85


def visualize_result(orig_mask, closed_mask, ellipse, ellipse_adj, actual_area, save_path, a, b, angle, major_axis_len,
                     minor_axis_len, error_percent, ellipse_area):
    """
    可视化结果并在图片左上角添加预测结果文本
    """
    vis = cv2.cvtColor(orig_mask, cv2.COLOR_GRAY2BGR)
    (cx, cy), (w_raw, h_raw), angle = ellipse
    (_, _), (w_adj, h_adj), _ = ellipse_adj

    COLOR_RAW = (0, 0, 255)
    COLOR_ADJ = (0, 255, 0)
    COLOR_MAJOR = (255, 165, 0)
    COLOR_MINOR = (123, 104, 238)
    COLOR_CENTER = (255, 0, 0)

    # 绘制椭圆和轴线
    cv2.ellipse(vis, ellipse, COLOR_RAW, thickness=2, lineType=cv2.LINE_AA)
    cv2.ellipse(vis, ellipse_adj, COLOR_ADJ, thickness=3, lineType=cv2.LINE_AA)

    major_angle_rad = math.radians(angle)
    major_end1 = (int(cx + a * math.cos(major_angle_rad)), int(cy + a * math.sin(major_angle_rad)))
    major_end2 = (int(cx - a * math.cos(major_angle_rad)), int(cy - a * math.sin(major_angle_rad)))
    cv2.line(vis, major_end1, major_end2, COLOR_MAJOR, 3, lineType=cv2.LINE_AA)

    minor_angle_rad = math.radians(angle + 90)
    minor_end1 = (int(cx + b * math.cos(minor_angle_rad)), int(cy + b * math.sin(minor_angle_rad)))
    minor_end2 = (int(cx - b * math.cos(minor_angle_rad)), int(cy - b * math.sin(minor_angle_rad)))
    cv2.line(vis, minor_end1, minor_end2, COLOR_MINOR, 3, lineType=cv2.LINE_AA)

    cv2.circle(vis, (int(cx), int(cy)), 5, COLOR_CENTER, -1, lineType=cv2.LINE_AA)

    # 在图片左上角添加预测结果文本
    font_face = cv2.FONT_HERSHEY_SIMPLEX
    font_scale = 1.2
    font_color = (255, 255, 255)
    font_thickness = 3
    line_type = cv2.LINE_AA

    # 文本背景色
    text_lines = [
        f"Major Axis: {major_axis_len}px",
        f"Minor Axis: {minor_axis_len}px",
        f"Area: {ellipse_area:.1f}px²",
        f"Error: {error_percent:.2f}%",
        f"Angle: {angle:.1f}°"
    ]

    # 计算文本区域大小
    text_height = 0
    max_width = 0
    for line in text_lines:
        text_size = cv2.getTextSize(line, font_face, font_scale, font_thickness)[0]
        max_width = max(max_width, text_size[0])
        text_height += text_size[1] + 10

    # 添加半透明背景
    bg_height = text_height + 20
    bg_width = max_width + 40
    overlay = vis.copy()
    cv2.rectangle(overlay, (10, 10), (bg_width, bg_height), (0, 0, 0), -1)
    cv2.addWeighted(overlay, 0.6, vis, 0.4, 0, vis)

    # 添加文本
    y_offset = 50
    for i, line in enumerate(text_lines):
        y = y_offset + i * 40
        cv2.putText(vis, line, (20, y), font_face, font_scale, font_color, font_thickness, line_type)

    # 保存图片
    if save_path.endswith('.png'):
        cv2.imwrite(save_path, vis, [cv2.IMWRITE_PNG_COMPRESSION, 0])
    else:
        cv2.imwrite(save_path, vis, [cv2.IMWRITE_JPEG_QUALITY, 95])
    print(f"  可视化结果已保存至: {save_path}")

    return vis


if __name__ == "__main__":
    # 使用示例：

    # 设置标定参数（根据实际情况调整）
    PX_AREA =  75.75 / 2048 / 1536  # 像素面积标定系数 (mm²/px)
    PX_LONG = np.sqrt(PX_AREA)  # 像素长度标定系数 (mm/px)
    H_THICKNESS = 5.0  # 板栗厚度 (mm)

    # 批量处理整个文件夹
    input_folder = "output_masks"
    all_results, features_data = batch_folder_predict(
        input_folder,
        ".",
        excel_output="板栗特征统计.xlsx",
        px_area=PX_AREA,
        px_long=PX_LONG,
        h=H_THICKNESS
    )

    # 显示处理结果摘要
    if features_data:
        print(f"\n✓ 成功处理 {len(features_data)} 个文件")
        print("导出的特征包括:")
        print("- 基本尺寸: 面积、长轴、短轴、厚度")
        print("- 形状参数: 形状索引、球形度、圆度比")
        print("- 几何特征: 算术/几何平均值、伸长率")
        print("- 物理参数: 表面积、体积")

        # 显示前几个文件的特征摘要
        print(f"\n前3个文件的特征预览:")
        for i, features in enumerate(features_data[:3]):
            print(f"{i + 1}. {features['文件名']}:")
            print(
                f"   面积: {features['标定面积']:.2f}, 长轴: {features['长半轴(a)']:.2f}, 短轴: {features['短半轴(b)']:.2f}")